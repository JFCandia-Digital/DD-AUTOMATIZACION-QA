#!/usr/bin/env python3
"""Genera matriz_api_test_recepcion_integracion_DocDigital.xlsx — QA-5687."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = "docs/api-v3.5/matriz_api_test_recepcion_integracion_DocDigital.xlsx"

HEADERS = [
    "ID",
    "Modulo",
    "Tipo Prueba",
    "Descripcion",
    "Precondiciones",
    "Pasos de Ejecucion",
    "Resultado Esperado",
    "Resultado Obtenido",
    "Estado",
    "Ambiente",
    "Evidencia",
    "Fecha",
    "Ejecutado Por",
    "Observaciones",
]

AMBIENTE = "QA - API v3.5"
EJECUTOR = "Juan Francisco Candia"
FECHA = "Pendiente"
TARJETA = "QA-5687"
ENDPOINT = "POST /pruebas-integracion/comunicaciones/test-recepcion"

BASE_PRECOND = (
    "VPN conectada; token OAuth válido de entidad integradora en QA; "
    "servidor QA con API_CLIENT_ENTIDAD_TEST=SEGPRES y API_SECRET_ENTIDAD_TEST configurados (Pablo); "
    ".env.api con credenciales de la entidad destinataria"
)

ROWS = [
    [
        "TC-INT-001",
        "API Integración",
        "API / Happy path",
        "Validar envío de comunicación de prueba con materia y folio (caso mínimo)",
        BASE_PRECOND,
        f"{ENDPOINT} | Body JSON: materia, folio (sin flags opcionales) | Validar status 200 | Validar result.id único",
        "200 OK; una comunicación creada; materia y folio reflejan parámetros de entrada; result.id numérico",
        "Pendiente ejecución — bloqueado 401/40100 por variables servidor (23-06)",
        "Pendiente",
        AMBIENTE,
        "Swagger / Cucumber @TestRecepcion_HappyPath; captura response JSON",
        FECHA,
        EJECUTOR,
        "CA: materia/folio entrada; creación individual (1 POST = 1 id)",
    ],
    [
        "TC-INT-002",
        "API Integración",
        "API / Funcional",
        "Validar envío con incorporaAnexos=true (anexos de prueba fijos del sistema)",
        BASE_PRECOND,
        f"{ENDPOINT} | incorporaAnexos: true | Validar 200 | GET detalle/anexos de la comunicación",
        "200 OK; comunicación incluye anexos generados por DD (mismos archivos de prueba siempre)",
        "Pendiente ejecución",
        "Pendiente",
        AMBIENTE,
        "Response JSON; GET comunicación o listado anexos",
        FECHA,
        EJECUTOR,
        "CA: flag opcional anexos; anexos siempre los mismos (validar en TC-INT-012)",
    ],
    [
        "TC-INT-003",
        "API Integración",
        "API / Funcional",
        "Validar envío con isReservado=true (documento principal marcado reservado)",
        BASE_PRECOND,
        f"{ENDPOINT} | isReservado: true | Validar 200 | Verificar flag reservado en detalle",
        "200 OK; comunicación creada con isReservado=true en metadatos",
        "Pendiente ejecución",
        "Pendiente",
        AMBIENTE,
        "Response JSON; detalle comunicación",
        FECHA,
        EJECUTOR,
        "CA: flag opcional reservado",
    ],
    [
        "TC-INT-004",
        "API Integración",
        "API / Funcional",
        "Validar envío con asociarProcedimientoAdministrativo=true",
        BASE_PRECOND,
        f"{ENDPOINT} | asociarProcedimientoAdministrativo: true | Validar 200 | GET detalle PA",
        "200 OK; PA asociado = primer PA de prueba 'Otro procedimiento'",
        "Pendiente ejecución",
        "Pendiente",
        AMBIENTE,
        "Response JSON; detalle comunicación / trazabilidad",
        FECHA,
        EJECUTOR,
        "CA: flag opcional PA; PA por defecto 'Otro procedimiento' (validar en TC-INT-014)",
    ],
    [
        "TC-INT-005",
        "API Integración",
        "API / Funcional",
        "Validar envío con todos los flags opcionales en true",
        BASE_PRECOND,
        f"{ENDPOINT} | incorporaAnexos, isReservado, asociarProcedimientoAdministrativo: true | Validar 200",
        "200 OK; comunicación con anexos, reservado y PA de prueba simultáneamente",
        "Pendiente ejecución",
        "Pendiente",
        AMBIENTE,
        "Swagger / Cucumber @TestRecepcion_FlagsCombinados",
        FECHA,
        EJECUTOR,
        "CA: combinación de los tres flags opcionales",
    ],
    [
        "TC-INT-006",
        "API Integración",
        "API / Validación",
        "Validar rechazo cuando falta materia en el body",
        BASE_PRECOND,
        f"{ENDPOINT} | Body sin campo materia (folio presente) | Validar 400",
        "400 Bad Request; mensaje de validación por materia obligatoria",
        "Pendiente ejecución",
        "Pendiente",
        AMBIENTE,
        "Swagger / Cucumber @TestRecepcion_Validacion",
        FECHA,
        EJECUTOR,
        "CA: materia es parámetro de entrada obligatorio",
    ],
    [
        "TC-INT-007",
        "API Integración",
        "API / Validación",
        "Validar rechazo cuando falta folio en el body",
        BASE_PRECOND,
        f"{ENDPOINT} | Body sin campo folio (materia presente) | Validar 400",
        "400 Bad Request; mensaje de validación por folio obligatorio",
        "Pendiente ejecución",
        "Pendiente",
        AMBIENTE,
        "Swagger / Cucumber @TestRecepcion_Validacion",
        FECHA,
        EJECUTOR,
        "CA: folio es parámetro de entrada obligatorio",
    ],
    [
        "TC-INT-008",
        "API Integración",
        "API / Seguridad",
        "Validar rechazo con token inválido o ausente",
        BASE_PRECOND.replace("token OAuth válido", "sin token o token inválido"),
        f"{ENDPOINT} | Authorization inválida o omitida | Validar 401",
        "401 Unauthorized; mismo criterio de seguridad que endpoints existentes",
        "401 — errorCode 40100 (23-06, antes de config servidor)",
        "Pendiente",
        AMBIENTE,
        "Captura Swagger 401; cucumber token inválido",
        FECHA,
        EJECUTOR,
        "CA: mismos criterios de seguridad. Nota: 401 previo también por ENV servidor",
    ],
    [
        "TC-INT-009",
        "API Integración",
        "API / Seguridad",
        "Validar rechazo con token expirado",
        BASE_PRECOND.replace("token OAuth válido", "token expirado"),
        f"{ENDPOINT} | Token expirado | Validar 401",
        "401 Unauthorized; mensaje sesión expirada (alineado a otros endpoints)",
        "Pendiente ejecución",
        "Pendiente",
        AMBIENTE,
        "Cucumber token expirado (patrón @Auth / otros features)",
        FECHA,
        EJECUTOR,
        "CA: mismos criterios de seguridad",
    ],
    [
        "TC-INT-010",
        "API Integración",
        "API / Contrato",
        "Validar que el integrador no puede elegir documento principal (solo JSON, sin multipart)",
        BASE_PRECOND,
        f"Revisar contrato Swagger {ENDPOINT} | Intentar envío multipart con PDF (si API lo rechaza) o confirmar que no existe campo archivo en spec",
        "Endpoint acepta solo JSON; no hay parámetro para documento principal; PDF lo fija el sistema (prueba_dipres_test.pdf con FEA)",
        "Pendiente ejecución",
        "Pendiente",
        AMBIENTE,
        "Captura Swagger spec; intento POST multipart documentado",
        FECHA,
        EJECUTOR,
        "CA: usuario no elige documento principal",
    ],
    [
        "TC-INT-011",
        "API Integración",
        "API / Negocio",
        "Validar que el documento principal es el mismo con isReservado true vs false",
        BASE_PRECOND,
        f"2x {ENDPOINT} misma materia/folio distinto timestamp | Post A: isReservado false | Post B: isReservado true | GET metadata/hash archivo principal de ambas",
        "Mismo archivo principal (hash/nombre/metadata idénticos); solo cambia flag reservado en metadatos",
        "Pendiente ejecución",
        "Pendiente",
        AMBIENTE,
        "GET archivo contenido/metadata; comparación hashes",
        FECHA,
        EJECUTOR,
        "CA: documento principal no cambia si es o no reservado",
    ],
    [
        "TC-INT-012",
        "API Integración",
        "API / Negocio",
        "Validar que los anexos son siempre los mismos entre envíos con incorporaAnexos=true",
        BASE_PRECOND,
        f"2x {ENDPOINT} con incorporaAnexos: true | GET listado anexos de ambas comunicaciones | Comparar nombres/hash",
        "Listado de anexos idéntico en ambas comunicaciones (archivos fijos generados por DD)",
        "Pendiente ejecución",
        "Pendiente",
        AMBIENTE,
        "GET anexos; comparación",
        FECHA,
        EJECUTOR,
        "CA: anexos siempre los mismos",
    ],
    [
        "TC-INT-013",
        "API Integración",
        "API / Negocio",
        "Validar entidad creadora y despachadora = Agosto KE (login interno)",
        BASE_PRECOND,
        f"{ENDPOINT} happy path | GET /comunicaciones/{{id}} o trazabilidad | Verificar entidadDespachadora",
        "Entidad creadora/despachadora corresponde a Agosto KE (codificador según QA)",
        "Pendiente ejecución",
        "Pendiente",
        AMBIENTE,
        "GET trazabilidad o detalle comunicación",
        FECHA,
        EJECUTOR,
        "CA: entidad creadora y despachadora es Agosto KE; login interno SEGPRES en servidor",
    ],
    [
        "TC-INT-014",
        "API Integración",
        "API / Negocio",
        "Validar PA por defecto 'Otro procedimiento' cuando asociarProcedimientoAdministrativo=true",
        BASE_PRECOND,
        f"{ENDPOINT} | asociarProcedimientoAdministrativo: true | GET detalle PA de la comunicación",
        "tipoProcedimientoAdministrativo = primer PA de prueba 'Otro procedimiento'",
        "Pendiente ejecución",
        "Pendiente",
        AMBIENTE,
        "Detalle comunicación JSON",
        FECHA,
        EJECUTOR,
        "CA: si asociado a PA se marca 'Otro procedimiento'",
    ],
    [
        "TC-INT-015",
        "API Integración",
        "API / Negocio",
        "Validar documento principal fijo con FEA en todos los envíos exitosos",
        BASE_PRECOND,
        f"{ENDPOINT} happy path | GET metadata/contenido documento principal | Verificar FEA y nombre fijo (prueba_dipres_test.pdf)",
        "Documento principal siempre el mismo PDF con firma electrónica válida generado por DD",
        "Pendiente ejecución",
        "Pendiente",
        AMBIENTE,
        "GET archivo metadata; captura propiedades FEA",
        FECHA,
        EJECUTOR,
        "CA: documento principal siempre el mismo con FEA",
    ],
    [
        "TC-INT-016",
        "API Integración",
        "API / Funcional",
        "Validar creación individual (no masiva): un POST genera exactamente una comunicación",
        BASE_PRECOND,
        f"{ENDPOINT} | Un único POST | Validar un result.id | Repetir POST distinto folio → nuevo id distinto",
        "Cada request crea una sola comunicación; no existe respuesta batch ni múltiples ids",
        "Pendiente ejecución",
        "Pendiente",
        AMBIENTE,
        "Dos responses JSON con ids distintos",
        FECHA,
        EJECUTOR,
        "CA: creación individual (no masiva)",
    ],
    [
        "TC-INT-017",
        "API Integración",
        "API / End-to-End",
        "Flujo integrador: envío prueba → pendientes recepción → acuse de recibo",
        BASE_PRECOND,
        f"{ENDPOINT} 200 | GET /comunicaciones/pendientes-recepcion?materia=... | PUT /comunicaciones/{{id}}/recepcion/{{tareaId}}/acuse-recibo",
        "POST 200; comunicación aparece en pendientes; acuse 200; flujo recepción completado",
        "Pendiente ejecución",
        "Pendiente",
        AMBIENTE,
        "Cucumber E2E; reports/index.html; features pendientes + acuse existentes",
        FECHA,
        EJECUTOR,
        "User story: probar método de recepción",
    ],
    [
        "TC-INT-018",
        "API Integración",
        "API / End-to-End",
        "Flujo integrador: envío prueba → pendientes recepción → rechazo",
        BASE_PRECOND,
        f"{ENDPOINT} 200 | GET pendientes-recepcion | PUT acuse-recibo con decisión rechazo",
        "POST 200; comunicación en pendientes; rechazo 200; flujo rechazo completado",
        "Pendiente ejecución",
        "Pendiente",
        AMBIENTE,
        "Cucumber E2E; putAcuseRecibo.feature",
        FECHA,
        EJECUTOR,
        "User story: probar método de rechazo",
    ],
]


def style_header(cell):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E79")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Matriz Test Recepción"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        style_header(cell)
        cell.border = border

    for r_idx, row_data in enumerate(ROWS, 2):
        for c_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if c_idx == 9:
                if value == "PASS":
                    cell.fill = PatternFill("solid", fgColor="C6EFCE")
                    cell.font = Font(color="006100", bold=True)
                elif value == "Pendiente":
                    cell.fill = PatternFill("solid", fgColor="FFEB9C")
                    cell.font = Font(color="9C6500", bold=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(ROWS) + 1}"

    widths = [14, 18, 18, 42, 38, 48, 38, 38, 12, 16, 36, 12, 20, 40]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    summary_row = len(ROWS) + 3
    ws.cell(row=summary_row, column=1, value="Tarjeta Asana").font = Font(bold=True)
    ws.cell(row=summary_row, column=2, value=f"{TARJETA} — Envío comunicación prueba para integración")
    ws.cell(row=summary_row + 1, column=1, value="Endpoint").font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=2, value=ENDPOINT)
    ws.cell(row=summary_row + 2, column=1, value="Bloqueo actual").font = Font(bold=True)
    ws.cell(
        row=summary_row + 2,
        column=2,
        value="Pendiente config servidor QA: API_CLIENT_ENTIDAD_TEST / API_SECRET_ENTIDAD_TEST (Susana → Pablo)",
    )
    ws.cell(row=summary_row + 3, column=1, value="Total casos").font = Font(bold=True)
    ws.cell(row=summary_row + 3, column=2, value=f"{len(ROWS)} casos — cubren 100% criterios de aceptación")

    wb.save(OUTPUT)
    print(f"Generado: {OUTPUT} ({len(ROWS)} casos)")


if __name__ == "__main__":
    main()
