#!/usr/bin/env python3
"""Genera matriz_api_notificacion_DocDigital.xlsx — formato plantilla QA DocDigital."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = "docs/api-v3.5/matriz_api_notificacion_DocDigital.xlsx"

HEADERS = [
    "ID",
    "Modulo",
    "Tipo Prueba",
    "Descripcion",
    "Precondiciones",
    "Pasos de Ejecucion",
    "Resultado Esperado",
    "Resultado Obtenido",
    "Estado",
    "Ambiente",
    "Evidencia",
    "Fecha",
    "Ejecutado Por",
    "Observaciones",
]

AMBIENTE = "QA - API v3.5"
EJECUTOR = "Juan Francisco Candia"
FECHA = "23-06-2026"
BASE_PRECOND = "VPN conectada, token PDI válido (CLIENT_ID_PDI), endpoint disponible, .env.api configurado"

ROWS = [
    [
        "TC-NOTIF-001",
        "API Notificación",
        "API / Integración",
        "Validar creación de comunicación tipo Notificación con documento firmado y datos PA válidos",
        f"{BASE_PRECOND}; entidad despachadora 598; destinatario 156; PDF 2_FIRMANTES_EN_DOC_DIGITAL.pdf",
        "POST /comunicaciones/despachar-tipo-notificacion (multipart) | Body: JSON_NOTIFICACION_HAPPY_PATH | Adjuntar PDF firmado | Validar status 200 | Validar result.id y result.fechaDespacho",
        "200 OK, comunicación creada con result.id y fechaDespacho actual",
        "200 OK — id 89703, fechaDespacho 23-06-2026 13:54:08; correo DocDigital 89703",
        "PASS",
        AMBIENTE,
        "reports/index.html; cucumber @Notificacion_HappyPath; correo DocDigital id 89703 (JFC)",
        FECHA,
        EJECUTOR,
        "GET pendientes + PUT acuse no automatizados en esta tarjeta; correo E2E de recepción recibido",
    ],
    [
        "TC-NOTIF-002",
        "API Notificación",
        "API / Validación",
        "Validar error al enviar request sin procedimiento administrativo",
        f"{BASE_PRECOND}; payload JSON_NOTIFICACION_SIN_TIPO_PROCEDIMIENTO",
        "POST /comunicaciones/despachar-tipo-notificacion | Body sin tipoProcedimientoAdministrativo | PDF sin firma (Prueba.pdf) | Validar status 400",
        "400 Bad Request, errorCode 40000, mensaje tipoProcedimientoAdministrativo es obligatorio",
        "400 — errorCode 40000 — tipoProcedimientoAdministrativo es obligatorio",
        "PASS",
        AMBIENTE,
        "reports/index.html; cucumber @Notificacion_CamposPA",
        FECHA,
        EJECUTOR,
        "",
    ],
    [
        "TC-NOTIF-003",
        "API Notificación",
        "API / End-to-End",
        "Validar flujo de notificación: despacho exitoso y recepción notificada en DocDigital",
        f"{BASE_PRECOND}; datos válidos despachadora 598; PDF firmado",
        "POST /comunicaciones/despachar-tipo-notificacion | Validar 200 + result.id | Verificar correo no-reply DocDigital con id de comunicación",
        "POST 200 OK; notificación de recepción enviada al destinatario (correo DocDigital)",
        "POST 200 OK — correos recibidos: 89703, 89705 (QA Notificacion JFC 13:54)",
        "PASS",
        AMBIENTE,
        "Captura bandeja correos 89703 y 89705; reports/index.html; cucumber suite 11/11",
        FECHA,
        EJECUTOR,
        "GET pendientes-recepción y PUT acuse-recibo no automatizados; evidencia E2E vía correo",
    ],
    [
        "TC-NOTIF-004",
        "API Notificación",
        "API / Validación",
        "Validar rechazo de notificación con documento principal sin firma digital",
        f"{BASE_PRECOND}; payload JSON_NOTIFICACION_VALIDO; PDF Prueba.pdf sin FEA",
        "POST /comunicaciones/despachar-tipo-notificacion | Adjuntar PDF sin firma | Validar status 400 y errorCode 4001",
        "400 Bad Request, errorCode 4001, mensaje archivo principal sin firmas externas",
        "400 — errorCode 4001 — Error en archivo principal sin firmas externas",
        "PASS",
        AMBIENTE,
        "reports/index.html; cucumber @Notificacion_SinFirma",
        FECHA,
        EJECUTOR,
        "",
    ],
    [
        "TC-NOTIF-005",
        "API Notificación",
        "API / Validación",
        "Validar error al enviar request sin destinatarios",
        f"{BASE_PRECOND}; payload JSON_NOTIFICACION_SIN_DESTINATARIOS",
        "POST /comunicaciones/despachar-tipo-notificacion | Body sin lista destinatarios | Validar 400 / 40000",
        "400 Bad Request, errorCode 40000, Se requiere al menos una entidad destinataria",
        "400 — errorCode 40000 — Se requiere al menos una entidad destinataria",
        "PASS",
        AMBIENTE,
        "reports/index.html; cucumber @Notificacion_CamposPA",
        FECHA,
        EJECUTOR,
        "",
    ],
    [
        "TC-NOTIF-006",
        "API Notificación",
        "API / Validación",
        "Validar error al enviar request sin configuracionDestinatarios",
        f"{BASE_PRECOND}; payload JSON_NOTIFICACION_SIN_CONFIGURACION_DESTINATARIOS",
        "POST /comunicaciones/despachar-tipo-notificacion | Body sin configuracionDestinatarios | Validar 400 / 40000",
        "400 Bad Request, errorCode 40000, configuracionDestinatarios es obligatorio",
        "400 — errorCode 40000 — configuracionDestinatarios es obligatorio",
        "PASS",
        AMBIENTE,
        "reports/index.html; cucumber @Notificacion_CamposPA",
        FECHA,
        EJECUTOR,
        "",
    ],
    [
        "TC-NOTIF-007",
        "API Notificación",
        "API / Validación",
        "Validar error al enviar request sin usuario solicitante",
        f"{BASE_PRECOND}; payload JSON_NOTIFICACION_SIN_USUARIO_SOLICITANTE",
        "POST /comunicaciones/despachar-tipo-notificacion | Body sin usuarioSolicitante | Validar 400 / 40000",
        "400 Bad Request, errorCode 40000, Usuario solicitante es obligatorio",
        "400 — errorCode 40000 — Usuario solicitante es obligatorio",
        "PASS",
        AMBIENTE,
        "reports/index.html; cucumber @Notificacion_CamposPA",
        FECHA,
        EJECUTOR,
        "",
    ],
    [
        "TC-NOTIF-008",
        "API Notificación",
        "API / Validación",
        "Validar error al enviar request sin RUN del usuario solicitante",
        f"{BASE_PRECOND}; payload JSON_NOTIFICACION_SIN_RUN_USUARIO_SOLICITANTE",
        "POST /comunicaciones/despachar-tipo-notificacion | Body sin run | Validar 400 / 40000",
        "400 Bad Request, errorCode 40000, RUN del usuario es obligatorio",
        "400 — errorCode 40000 — RUN del usuario es obligatorio",
        "PASS",
        AMBIENTE,
        "reports/index.html; cucumber @Notificacion_CamposPA",
        FECHA,
        EJECUTOR,
        "",
    ],
    [
        "TC-NOTIF-009",
        "API Notificación",
        "API / Validación",
        "Validar error al enviar request sin dígito verificador (DV) del RUN",
        f"{BASE_PRECOND}; payload JSON_NOTIFICACION_SIN_DV_USUARIO_SOLICITANTE",
        "POST /comunicaciones/despachar-tipo-notificacion | Body sin dv | Validar 400 / 40000",
        "400 Bad Request, errorCode 40000, Digito verificador del RUN es obligatorio",
        "400 — errorCode 40000 — Digito verificador del RUN es obligatorio",
        "PASS",
        AMBIENTE,
        "reports/index.html; cucumber @Notificacion_CamposPA",
        FECHA,
        EJECUTOR,
        "",
    ],
    [
        "TC-NOTIF-010",
        "API Notificación",
        "API / Negocio",
        "Validar despacho con entidad dependiente (4758) como destinatario en copia",
        f"{BASE_PRECOND}; despachadora 598; dependiente 4758 (GET /entidades/dependientes/598); principal 156",
        "POST /comunicaciones/despachar-tipo-notificacion | JSON_NOTIFICACION_DEPENDIENTE_EN_COPIA | PDF firmado | Validar 200 + id",
        "200 OK, comunicación creada, dependiente en copia aceptado",
        "200 OK — correo DocDigital 89705 (23-06-2026 13:54:17); materia QA Notificacion JFC",
        "PASS",
        AMBIENTE,
        "reports/index.html; cucumber @Notificacion_Dependientes; captura correo",
        FECHA,
        EJECUTOR,
        "",
    ],
    [
        "TC-NOTIF-011",
        "API Notificación",
        "API / Negocio",
        "Validar despacho con entidad dependiente (4758) como destinatario principal",
        f"{BASE_PRECOND}; despachadora 598; dependiente 4758 como único destinatario principal",
        "POST /comunicaciones/despachar-tipo-notificacion | JSON_NOTIFICACION_DEPENDIENTE_COMO_PRINCIPAL | PDF firmado | Validar respuesta API",
        "Según manual: rechazo 400; en QA observado: 200 OK con id generado",
        "200 OK — API QA acepta dependiente como principal (ver hallazgo H-01 en evidencias)",
        "PASS",
        AMBIENTE,
        "reports/index.html; cucumber @Notificacion_Dependientes",
        FECHA,
        EJECUTOR,
        "Hallazgo H-01: API QA acepta dependiente como principal; confirmar regla con negocio/backend",
    ],
    [
        "TC-NOTIF-012",
        "API Notificación",
        "API / Negocio",
        "Validar rechazo con destinatario no dependiente de la entidad despachadora",
        f"{BASE_PRECOND}; destinatario no dependiente (598); payload JSON_NOTIFICACION_DESTINATARIO_NO_DEPENDIENTE",
        "POST /comunicaciones/despachar-tipo-notificacion | PDF sin firma | Validar 400",
        "400 Bad Request por regla de negocio o validación de documento",
        "400 — errorCode 4001 (validación firma alcanzada antes que regla dependiente)",
        "PASS",
        AMBIENTE,
        "reports/index.html; cucumber @Notificacion_Dependientes",
        FECHA,
        EJECUTOR,
        "Hallazgo H-02: cobertura parcial de regla dependiente; usar PDF firmado para profundizar en otra iteración",
    ],
]


def style_header(cell):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E79")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Matriz Notificación"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        style_header(cell)
        cell.border = border

    for r_idx, row_data in enumerate(ROWS, 2):
        for c_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if c_idx == 9:
                if value == "PASS":
                    cell.fill = PatternFill("solid", fgColor="C6EFCE")
                    cell.font = Font(color="006100", bold=True)
                elif value == "Pendiente":
                    cell.fill = PatternFill("solid", fgColor="FFEB9C")
                    cell.font = Font(color="9C6500")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(ROWS) + 1}"

    widths = [14, 18, 18, 42, 38, 48, 38, 38, 10, 16, 36, 12, 16, 40]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Fila resumen al final
    summary_row = len(ROWS) + 3
    ws.cell(row=summary_row, column=1, value="Resumen suite").font = Font(bold=True)
    ws.cell(
        row=summary_row,
        column=2,
        value="npm run notificacion_report → 11 scenarios (11 passed) | 12 casos matriz",
    )

    wb.save(OUTPUT)
    print(f"Generado: {OUTPUT} ({len(ROWS)} casos)")


if __name__ == "__main__":
    main()
